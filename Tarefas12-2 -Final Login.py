import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os

class Tarefa:
    def __init__(self, descricao, data, horario, disciplina, prioridade, responsavel, status="Pendente"):
        self.descricao = descricao
        self.data = data
        self.horario = horario
        self.disciplina = disciplina
        self.prioridade = prioridade
        self.status = status
        self.responsavel = responsavel

    def __str__(self):
        return f"{self.descricao} - {self.disciplina} - {self.status} - {self.data} {self.horario} - Prioridade: {self.prioridade}"

class GerenciadorDeTarefas:
    def __init__(self, master):
        self.master = master
        self.master.title("Gerenciador de Tarefas - Agenda")
        self.master.geometry("600x450")
        self.master.config(bg="#f0f0f0")

        self.arquivo_excel = "tarefas.xlsx"
        if not os.path.exists(self.arquivo_excel):
            self.criar_arquivo_excel()

        title_label = ttk.Label(master, text="Gerenciador de Tarefas", font=("Arial", 16), background="#f0f0f0")
        title_label.pack(pady=10)

        # Frame para entrada de dados
        frame_entry = ttk.Frame(master)
        frame_entry.pack(pady=10)

        ttk.Label(frame_entry, text="Descrição:", background="#f0f0f0").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.descricao_entry = ttk.Entry(frame_entry, width=40)
        self.descricao_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame_entry, text="Data (DD-MM-AAAA):", background="#f0f0f0").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.data_entry = ttk.Entry(frame_entry, width=20)
        self.data_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(frame_entry, text="Horário (HH:MM):", background="#f0f0f0").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        self.horario_entry = ttk.Entry(frame_entry, width=20)
        self.horario_entry.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(frame_entry, text="Disciplina:", background="#f0f0f0").grid(row=3, column=0, padx=5, pady=5, sticky='w')
        self.disciplina_entry = ttk.Entry(frame_entry, width=40)
        self.disciplina_entry.grid(row=3, column=1, padx=5, pady=5)

        ttk.Label(frame_entry, text="Prioridade:", background="#f0f0f0").grid(row=4, column=0, padx=5, pady=5, sticky='w')
        self.prioridade_combobox = ttk.Combobox(frame_entry, values=["Alta", "Média", "Baixa"], width=18)
        self.prioridade_combobox.grid(row=4, column=1, padx=5, pady=5)

        ttk.Label(frame_entry, text="Responsável:", background="#f0f0f0").grid(row=5, column=0, padx=5, pady=5, sticky='w')
        self.responsavel_entry = ttk.Entry(frame_entry, width=40)
        self.responsavel_entry.grid(row=5, column=1, padx=5, pady=5)

        # Separador
        ttk.Separator(master, orient='horizontal').pack(fill='x', pady=10)

        # Botões de ação
        button_frame = ttk.Frame(master)
        button_frame.pack(pady=5)

        self.add_button = ttk.Button(button_frame, text="Adicionar Tarefa", command=self.adicionar_tarefa)
        self.add_button.grid(row=0, column=0, padx=5, pady=5)

        self.remove_button = ttk.Button(button_frame, text="Remover Tarefa", command=self.remover_tarefa)
        self.remove_button.grid(row=0, column=1, padx=5, pady=5)

        # Listbox com Scrollbar
        self.listbox_frame = ttk.Frame(master)
        self.listbox_frame.pack(pady=10)

        self.listbox = tk.Listbox(self.listbox_frame, width=80, height=10, bg="#ffffff", font=("Arial", 10))
        self.listbox.pack(side=tk.LEFT)

        self.scrollbar = ttk.Scrollbar(self.listbox_frame, orient=tk.VERTICAL, command=self.listbox.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.listbox.config(yscrollcommand=self.scrollbar.set)

        # Carregar as tarefas do arquivo Excel
        self.carregar_tarefas_excel()

    def criar_arquivo_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tarefas"
        ws.append(["Descrição", "Data", "Horário", "Disciplina", "Prioridade", "Status", "Responsável"])
        wb.save(self.arquivo_excel)

    def adicionar_tarefa(self):
        descricao = self.descricao_entry.get()
        data = self.data_entry.get()
        horario = self.horario_entry.get()
        disciplina = self.disciplina_entry.get()
        prioridade = self.prioridade_combobox.get()
        responsavel = self.responsavel_entry.get()

        if descricao and data and horario and disciplina and prioridade and responsavel:
            nova_tarefa = Tarefa(descricao, data, horario, disciplina, prioridade, responsavel)
            self.salvar_tarefa_excel(nova_tarefa)
            self.carregar_tarefas_excel()  # Atualiza a lista com a nova tarefa
            self.limpar_campos()
            messagebox.showinfo("Sucesso", "Tarefa adicionada com sucesso!")
        else:
            messagebox.showwarning("Erro", "Todos os campos devem ser preenchidos.")

    def salvar_tarefa_excel(self, tarefa):
        wb = load_workbook(self.arquivo_excel)
        ws = wb.active
        ws.append([tarefa.descricao, tarefa.data, tarefa.horario, tarefa.disciplina, tarefa.prioridade, tarefa.status, tarefa.responsavel])
        wb.save(self.arquivo_excel)

    def carregar_tarefas_excel(self):
        self.listbox.delete(0, tk.END)
        wb = load_workbook(self.arquivo_excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # Ignora a primeira linha de cabeçalhos
            tarefa_obj = Tarefa(*row)
            self.listbox.insert(tk.END, tarefa_obj)

    def limpar_campos(self):
        self.descricao_entry.delete(0, tk.END)
        self.data_entry.delete(0, tk.END)
        self.horario_entry.delete(0, tk.END)
        self.disciplina_entry.delete(0, tk.END)
        self.prioridade_combobox.set('')  # Limpa a seleção do Combobox
        self.responsavel_entry.delete(0, tk.END)

    def remover_tarefa(self):
        try:
            index = self.listbox.curselection()[0]
            tarefa_texto = self.listbox.get(index)
            descricao_tarefa = tarefa_texto.split(" - ")[0]  # Extrai a descrição da tarefa
            self.remover_tarefa_excel(descricao_tarefa)
            self.carregar_tarefas_excel()  # Atualiza a lista após a remoção
            messagebox.showinfo("Sucesso", "Tarefa removida com sucesso!")
        except IndexError:
            messagebox.showwarning("Erro", "Selecione uma tarefa para remover.")

    def remover_tarefa_excel(self, descricao):
        wb = load_workbook(self.arquivo_excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == descricao:
                ws.delete_rows(row[0].row, 1)
                break
        wb.save(self.arquivo_excel)

class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("Login")
        self.root.geometry("300x200")

        # Label e campos de entrada para Login
        self.label_username = ttk.Label(root, text="Usuário:")
        self.label_username.pack(pady=5)
        self.entry_username = ttk.Entry(root, width=30)
        self.entry_username.pack(pady=5)

        self.label_password = ttk.Label(root, text="Senha:")
        self.label_password.pack(pady=5)
        self.entry_password = ttk.Entry(root, width=30, show="*")
        self.entry_password.pack(pady=5)

        self.login_button = ttk.Button(root, text="Entrar", command=self.login)
        self.login_button.pack(pady=10)

    def login(self):
        username = self.entry_username.get()
        password = self.entry_password.get()

        # Aqui você pode substituir por um banco de dados ou arquivo para validação real
        usuarios_validos = {
            "admin": "senha123",  # Exemplo de usuário e senha
            "usuario1": "12345"
        }

        if usuarios_validos.get(username) == password:
            messagebox.showinfo("Sucesso", "Login realizado com sucesso!")
            self.root.destroy()  # Fecha a janela de login
            self.abrir_gerenciador_tarefas()
        else:
            messagebox.showerror("Erro", "Usuário ou senha inválidos!")

    def abrir_gerenciador_tarefas(self):
        root_tarefas = tk.Tk()
        app = GerenciadorDeTarefas(root_tarefas)
        root_tarefas.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    login_app = LoginWindow(root)
    root.mainloop()
